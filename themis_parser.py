from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SUMMARY_SHEET = "Tổng hợp điểm"
DETAIL_SHEET = "Chi tiết chấm"
SUMMARY_REQUIRED = {"Mã thí sinh", "Tổng điểm"}
DETAIL_REQUIRED = {"Mã thí sinh", "Bài thi", "Test", "Điểm", "Ghi chú"}


class WorkbookValidationError(ValueError):
    pass


@dataclass
class ParsedWorkbook:
    room: str
    problems: list[str]
    contestants: list[dict[str, Any]]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).replace("_x000D_", "\n").strip()


def clean_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def detect_verdict(note: str, score: float) -> str:
    lower = note.lower()
    if "dịch thành công" in lower or "compile" in lower and "thành công" in lower:
        return "Compile"
    if "dịch lỗi" in lower or "compilation" in lower and "error" in lower:
        return "CE"
    if "quá thời gian" in lower or "time limit" in lower or "tle" in lower:
        return "TLE"
    if "runtime" in lower or "run time" in lower or "lỗi khi chạy" in lower:
        return "RTE"
    if "khớp đáp án" in lower or "accepted" in lower or "correct" in lower:
        return "AC"
    if "sai" in lower or "không khớp" in lower or "khác đáp án" in lower or "wrong answer" in lower:
        return "WA"
    if score > 0:
        return "Partial"
    return "Unknown"


def parse_runtime_ms(note: str) -> float | None:
    match = re.search(r"Thời gian\s*[≈=]?\s*([0-9]+(?:[,.][0-9]+)?)\s*giây", note, re.IGNORECASE)
    if not match:
        match = re.search(r"([0-9]+(?:[,.][0-9]+)?)\s*s(?:ec(?:ond)?s?)?\b", note, re.IGNORECASE)
    if not match:
        return None
    seconds = float(match.group(1).replace(",", "."))
    return round(seconds * 1000, 3)


def short_note(note: str, max_len: int = 220) -> str:
    compact = re.sub(r"\s+", " ", note).strip()
    if len(compact) <= max_len:
        return compact
    return compact[: max_len - 1].rstrip() + "…"


def _load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise WorkbookValidationError(f"File {path.name} không phải Excel hợp lệ.") from exc
    except OSError as exc:
        raise WorkbookValidationError(f"Không đọc được file {path.name}.") from exc

    if sheet_name not in workbook.sheetnames:
        raise WorkbookValidationError(f"Thiếu sheet '{sheet_name}' trong file {path.name}.")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise WorkbookValidationError(f"Sheet '{sheet_name}' trong file {path.name} đang trống.") from exc

    headers = [clean_text(value) for value in header_row]
    data: list[dict[str, Any]] = []
    for row in rows:
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        if any(clean_text(value) for value in item.values()):
            data.append(item)
    workbook.close()
    return data


def _validate_columns(rows: list[dict[str, Any]], required: set[str], sheet_name: str, filename: str) -> None:
    columns = set(rows[0].keys()) if rows else set()
    missing = sorted(required.difference(columns))
    if missing:
        raise WorkbookValidationError(
            f"File {filename}, sheet '{sheet_name}' thiếu cột: {', '.join(missing)}."
        )


def parse_workbook(path: str | Path, room: str) -> ParsedWorkbook:
    file_path = Path(path)
    summary_rows = _load_rows(file_path, SUMMARY_SHEET)
    detail_rows = _load_rows(file_path, DETAIL_SHEET)
    _validate_columns(summary_rows, SUMMARY_REQUIRED, SUMMARY_SHEET, file_path.name)
    _validate_columns(detail_rows, DETAIL_REQUIRED, DETAIL_SHEET, file_path.name)

    columns = list(summary_rows[0].keys())
    total_index = columns.index("Tổng điểm")
    code_index = columns.index("Mã thí sinh")
    if total_index <= code_index + 1:
        raise WorkbookValidationError(f"File {file_path.name} không có cột bài thi.")
    problems = [str(col).strip() for col in columns[code_index + 1 : total_index] if str(col).strip()]

    detail_by_code: dict[str, list[dict[str, Any]]] = {}
    for row in detail_rows:
        code = clean_text(row.get("Mã thí sinh"))
        problem_code = clean_text(row.get("Bài thi"))
        test_name = clean_text(row.get("Test"))
        if not code or not problem_code or not test_name:
            continue
        score = clean_float(row.get("Điểm"))
        note = clean_text(row.get("Ghi chú"))
        detail_by_code.setdefault(code, []).append(
            {
                "problem_code": problem_code,
                "test_name": test_name,
                "score": score,
                "verdict": detect_verdict(note, score),
                "runtime_ms": parse_runtime_ms(note),
                "note": short_note(note),
            }
        )

    contestants: list[dict[str, Any]] = []
    for row in summary_rows:
        code = clean_text(row.get("Mã thí sinh"))
        if not code:
            continue
        problem_scores = {problem: clean_float(row.get(problem)) for problem in problems}
        contestants.append(
            {
                "room": room,
                "code": code,
                "total_score": clean_float(row.get("Tổng điểm")),
                "problem_scores": problem_scores,
                "test_results": detail_by_code.get(code, []),
            }
        )

    return ParsedWorkbook(room=room, problems=problems, contestants=contestants)


def build_contest_payload(workbooks: list[ParsedWorkbook]) -> tuple[list[str], dict[str, Any], list[dict[str, Any]]]:
    problems: list[str] = []
    seen = set()
    contestants: list[dict[str, Any]] = []
    for workbook in workbooks:
        for problem in workbook.problems:
            if problem not in seen:
                seen.add(problem)
                problems.append(problem)
        contestants.extend(workbook.contestants)

    contestants.sort(key=lambda item: (-item["total_score"], item["code"].casefold(), item["room"].casefold()))
    last_score: float | None = None
    current_rank = 0
    for index, contestant in enumerate(contestants, start=1):
        if last_score is None or contestant["total_score"] != last_score:
            current_rank = index
            last_score = contestant["total_score"]
        contestant["rank"] = current_rank
        contestant["is_top_35"] = index <= 35
        for problem in problems:
            contestant["problem_scores"].setdefault(problem, 0.0)

    count = len(contestants)
    total_avg = round(sum(item["total_score"] for item in contestants) / count, 3) if count else 0.0
    problem_averages = {}
    for problem in problems:
        problem_averages[problem] = (
            round(sum(item["problem_scores"].get(problem, 0.0) for item in contestants) / count, 3)
            if count
            else 0.0
        )

    stats = {
        "contestant_count": count,
        "total_average": total_avg,
        "problem_averages": problem_averages,
    }
    return problems, stats, contestants
